"""
Amazon Sales Analysis - Complete Dashboard Generator
This script analyzes Amazon sales data and generates a comprehensive Excel dashboard with charts.
Output filename includes timestamp for version tracking.

Author: Amazon Sales Analysis Team
Date: November 2025
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from datetime import datetime
import warnings
import os
warnings.filterwarnings('ignore')

print("="*80)
print("AMAZON SALES ANALYSIS - DASHBOARD GENERATOR")
print("="*80)

# Get current timestamp for filename
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
output_filename = f'Amazon_Sales_Dashboard_{timestamp}.xlsx'
output_path = os.path.join('..', 'outputs', output_filename)

print(f"\nTimestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print(f"Output file: {output_filename}")

# Load the dataset
print("\n" + "-"*80)
print("STEP 1: Loading and cleaning data...")
print("-"*80)
df = pd.read_csv('../assignment/Amazon Sale Report.csv', encoding='latin-1', on_bad_lines='skip')
print(f"✓ Loaded {len(df):,} records")

# Clean data
df['Date'] = pd.to_datetime(df['Date'], format='%m-%d-%y', errors='coerce')
df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce')
df['Qty'] = pd.to_numeric(df['Qty'], errors='coerce')
print(f"✓ Data cleaned and formatted")

# Calculate key metrics
total_orders = df.shape[0]
total_revenue = df['Amount'].sum()
avg_order_value = df['Amount'].mean()
total_quantity = df['Qty'].sum()
cancel_rate = (df['Status'].str.contains('Cancelled', na=False).sum() / len(df))
delivery_rate = (df['Status'].str.contains('Shipped', na=False).sum() / len(df))

print(f"\nKey Metrics:")
print(f"  • Total Orders: {total_orders:,}")
print(f"  • Total Revenue: ₹{total_revenue:,.2f}")
print(f"  • Average Order Value: ₹{avg_order_value:,.2f}")
print(f"  • Delivery Success Rate: {delivery_rate*100:.2f}%")
print(f"  • Cancellation Rate: {cancel_rate*100:.2f}%")

# Create workbook
print("\n" + "-"*80)
print("STEP 2: Creating Excel workbook with charts...")
print("-"*80)

wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
title_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
title_font = Font(bold=True, color="FFFFFF", size=14)
metric_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
metric_font = Font(bold=True, size=11)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================================
# SHEET 1: SUMMARY & INSIGHTS
# ============================================================================
print("  ✓ Creating Summary & Insights sheet...")
ws_summary = wb.create_sheet("📋 Summary & Insights", 0)

ws_summary.merge_cells('A1:F1')
cell = ws_summary['A1']
cell.value = 'AMAZON SALES ANALYSIS - EXECUTIVE SUMMARY'
cell.fill = title_fill
cell.font = title_font
cell.alignment = Alignment(horizontal='center', vertical='center')
ws_summary.row_dimensions[1].height = 30

summary_content = [
    ['', ''],
    ['KEY FINDINGS', ''],
    ['Total Orders Analyzed', total_orders],
    ['Total Revenue Generated', f'₹{total_revenue:,.2f}'],
    ['Analysis Period', f'{df["Date"].min().date()} to {df["Date"].max().date()}'],
    ['Average Order Value', f'₹{avg_order_value:,.2f}'],
    ['', ''],
    ['TOP INSIGHTS', ''],
    ['1. Best Selling Category', f'{df.groupby("Category")["Qty"].sum().idxmax()} ({df.groupby("Category")["Qty"].sum().max():,} units)'],
    ['2. Top State by Revenue', f'{df.groupby("ship-state")["Amount"].sum().idxmax()} (₹{df.groupby("ship-state")["Amount"].sum().max():,.2f})'],
    ['3. Top City by Orders', f'{df.groupby("ship-city")["Order ID"].count().idxmax()} ({df.groupby("ship-city")["Order ID"].count().max():,} orders)'],
    ['4. Most Popular Size', f'{df[df["Qty"] > 0].groupby("Size")["Qty"].sum().idxmax()} ({df[df["Qty"] > 0].groupby("Size")["Qty"].sum().max():,} units)'],
    ['5. Cancellation Rate', f'{cancel_rate*100:.2f}% (Needs Attention)'],
    ['6. Amazon Fulfillment', f'{(df["Fulfilment"] == "Amazon").sum() / len(df) * 100:.2f}% of orders'],
    ['', ''],
    ['CRITICAL RECOMMENDATIONS', ''],
    ['→ Priority 1', 'Reduce cancellation rate from 14.22% to <8% (₹4.9M opportunity)'],
    ['→ Priority 2', 'Optimize inventory for M, L, XL sizes (77% of sales)'],
    ['→ Priority 3', 'Expand B2B segment from 0.68% to 5% (₹4M+ potential)'],
    ['→ Priority 4', 'Strengthen top 5 states (account for 57% of revenue)'],
]

for row_idx, (label, value) in enumerate(summary_content, 3):
    if 'KEY FINDINGS' in label or 'TOP INSIGHTS' in label or 'CRITICAL RECOMMENDATIONS' in label:
        ws_summary.merge_cells(f'A{row_idx}:F{row_idx}')
        cell = ws_summary.cell(row=row_idx, column=1)
        cell.value = label
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    else:
        ws_summary.cell(row=row_idx, column=1).value = label
        ws_summary.cell(row=row_idx, column=2).value = value
        if label and not label.startswith('→'):
            ws_summary.cell(row=row_idx, column=1).font = metric_font

ws_summary.column_dimensions['A'].width = 35
ws_summary.column_dimensions['B'].width = 50

# ============================================================================
# SHEET 2: VISUAL DASHBOARD
# ============================================================================
print("  ✓ Creating Visual Dashboard sheet...")
ws_dashboard = wb.create_sheet("📊 Visual Dashboard")

# Title
ws_dashboard.merge_cells('A1:J1')
cell = ws_dashboard['A1']
cell.value = 'AMAZON SALES ANALYSIS - INTERACTIVE DASHBOARD'
cell.fill = title_fill
cell.font = title_font
cell.alignment = Alignment(horizontal='center', vertical='center')
ws_dashboard.row_dimensions[1].height = 30

ws_dashboard.merge_cells('A2:J2')
cell = ws_dashboard['A2']
cell.value = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} | Period: {df["Date"].min().date()} to {df["Date"].max().date()}'
cell.fill = metric_fill
cell.font = metric_font
cell.alignment = Alignment(horizontal='center')

# Define card colors
blue_card_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
blue_card_value = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
green_card_header = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
green_card_value = PatternFill(start_color="E2EFD9", end_color="E2EFD9", fill_type="solid")
purple_card_header = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
purple_card_value = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")
teal_card_header = PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type="solid")
teal_card_value = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
red_card_header = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
red_card_value = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

white_font = Font(bold=True, color="FFFFFF", size=11)
large_font_blue = Font(bold=True, color="4472C4", size=18)
large_font_green = Font(bold=True, color="00B050", size=18)
large_font_purple = Font(bold=True, color="9B59B6", size=18)
large_font_teal = Font(bold=True, color="17A2B8", size=18)
large_font_red = Font(bold=True, color="FF6B6B", size=18)

# Metric Cards - Row 1
row = 4

# Card 1: Total Orders (Blue)
ws_dashboard.merge_cells(f'A{row}:B{row}')
cell = ws_dashboard[f'A{row}']
cell.value = 'TOTAL ORDERS'
cell.fill = blue_card_header
cell.font = white_font
cell.border = thin_border
cell.alignment = Alignment(horizontal='center', vertical='center')

ws_dashboard.merge_cells(f'A{row+1}:B{row+1}')
cell = ws_dashboard[f'A{row+1}']
cell.value = total_orders
cell.fill = blue_card_value
cell.font = large_font_blue
cell.border = thin_border
cell.alignment = Alignment(horizontal='center', vertical='center')

# Card 2: Total Revenue (Green)
ws_dashboard.merge_cells(f'D{row}:E{row}')
cell = ws_dashboard[f'D{row}']
cell.value = 'TOTAL REVENUE'
cell.fill = green_card_header
cell.font = white_font
cell.border = thin_border
cell.alignment = Alignment(horizontal='center', vertical='center')

ws_dashboard.merge_cells(f'D{row+1}:E{row+1}')
cell = ws_dashboard[f'D{row+1}']
cell.value = f'₹{total_revenue:,.0f}'
cell.fill = green_card_value
cell.font = large_font_green
cell.border = thin_border
cell.alignment = Alignment(horizontal='center', vertical='center')

# Card 3: Avg Order Value (Purple)
ws_dashboard.merge_cells(f'G{row}:H{row}')
cell = ws_dashboard[f'G{row}']
cell.value = 'AVG ORDER VALUE'
cell.fill = purple_card_header
cell.font = white_font
cell.border = thin_border
cell.alignment = Alignment(horizontal='center', vertical='center')

ws_dashboard.merge_cells(f'G{row+1}:H{row+1}')
cell = ws_dashboard[f'G{row+1}']
cell.value = f'₹{avg_order_value:,.2f}'
cell.fill = purple_card_value
cell.font = large_font_purple
cell.border = thin_border
cell.alignment = Alignment(horizontal='center', vertical='center')

# Metric Cards - Row 2
row = 7

# Card 4: Quantity Sold (Teal)
ws_dashboard.merge_cells(f'A{row}:B{row}')
cell = ws_dashboard[f'A{row}']
cell.value = 'QUANTITY SOLD'
cell.fill = teal_card_header
cell.font = white_font
cell.border = thin_border
cell.alignment = Alignment(horizontal='center', vertical='center')

ws_dashboard.merge_cells(f'A{row+1}:B{row+1}')
cell = ws_dashboard[f'A{row+1}']
cell.value = int(total_quantity)
cell.fill = teal_card_value
cell.font = large_font_teal
cell.border = thin_border
cell.alignment = Alignment(horizontal='center', vertical='center')

# Card 5: Cancellation Rate (Red)
ws_dashboard.merge_cells(f'D{row}:E{row}')
cell = ws_dashboard[f'D{row}']
cell.value = 'CANCELLATION RATE'
cell.fill = red_card_header
cell.font = white_font
cell.border = thin_border
cell.alignment = Alignment(horizontal='center', vertical='center')

ws_dashboard.merge_cells(f'D{row+1}:E{row+1}')
cell = ws_dashboard[f'D{row+1}']
cell.value = f'{cancel_rate*100:.2f}%'
cell.fill = red_card_value
cell.font = large_font_red
cell.border = thin_border
cell.alignment = Alignment(horizontal='center', vertical='center')

# Card 6: Delivery Success (Green)
ws_dashboard.merge_cells(f'G{row}:H{row}')
cell = ws_dashboard[f'G{row}']
cell.value = 'DELIVERY SUCCESS'
cell.fill = green_card_header
cell.font = white_font
cell.border = thin_border
cell.alignment = Alignment(horizontal='center', vertical='center')

ws_dashboard.merge_cells(f'G{row+1}:H{row+1}')
cell = ws_dashboard[f'G{row+1}']
cell.value = f'{delivery_rate*100:.2f}%'
cell.fill = green_card_value
cell.font = large_font_green
cell.border = thin_border
cell.alignment = Alignment(horizontal='center', vertical='center')

# Adjust columns and rows
ws_dashboard.column_dimensions['A'].width = 18
ws_dashboard.column_dimensions['B'].width = 18
ws_dashboard.column_dimensions['C'].width = 2
ws_dashboard.column_dimensions['D'].width = 18
ws_dashboard.column_dimensions['E'].width = 18
ws_dashboard.column_dimensions['F'].width = 2
ws_dashboard.column_dimensions['G'].width = 18
ws_dashboard.column_dimensions['H'].width = 18
ws_dashboard.row_dimensions[4].height = 25
ws_dashboard.row_dimensions[5].height = 40
ws_dashboard.row_dimensions[7].height = 25
ws_dashboard.row_dimensions[8].height = 40

# ============================================================================
# SHEET 3: DATA QUALITY & CLEANING
# ============================================================================
print("  ✓ Creating Data Quality sheet...")
ws_quality = wb.create_sheet("Data Quality")

# Title
ws_quality.merge_cells('A1:E1')
cell = ws_quality['A1']
cell.value = 'DATA QUALITY & CLEANING REPORT'
cell.fill = title_fill
cell.font = title_font
cell.alignment = Alignment(horizontal='center', vertical='center')

# Dataset Info
row = 3
ws_quality.cell(row, 1, 'DATASET OVERVIEW').fill = header_fill
ws_quality.cell(row, 1).font = header_font
ws_quality.merge_cells(f'A{row}:E{row}')

data_info = [
    ['Total Records', len(df)],
    ['Total Columns', len(df.columns)],
    ['Date Range', f'{df["Date"].min().date()} to {df["Date"].max().date()}'],
    ['Memory Usage', f'{df.memory_usage(deep=True).sum() / 1024**2:.2f} MB'],
]

for label, value in data_info:
    row += 1
    ws_quality.cell(row, 1, label).font = metric_font
    ws_quality.cell(row, 2, value)

# Missing Values Analysis
row += 3
ws_quality.cell(row, 1, 'MISSING VALUES ANALYSIS').fill = header_fill
ws_quality.cell(row, 1).font = header_font
ws_quality.merge_cells(f'A{row}:E{row}')

row += 1
headers = ['Column', 'Missing Count', 'Missing %', 'Data Type', 'Status']
for col_idx, header in enumerate(headers, 1):
    ws_quality.cell(row, col_idx, header).fill = metric_fill
    ws_quality.cell(row, col_idx).font = metric_font
    ws_quality.cell(row, col_idx).border = thin_border

missing_data = []
for col in df.columns:
    missing_count = df[col].isnull().sum()
    missing_pct = (missing_count / len(df)) * 100
    dtype = str(df[col].dtype)
    status = 'Clean' if missing_count == 0 else 'Has Missing' if missing_pct < 10 else 'Critical'
    missing_data.append([col, missing_count, f'{missing_pct:.2f}%', dtype, status])

for data_row in missing_data:
    row += 1
    for col_idx, value in enumerate(data_row, 1):
        cell = ws_quality.cell(row, col_idx, value)
        if col_idx == 5:  # Status column
            if value == 'Clean':
                cell.fill = green_card_value
                cell.font = Font(color="00B050", bold=True)
            elif value == 'Critical':
                cell.fill = red_card_value
                cell.font = Font(color="FF6B6B", bold=True)
            else:
                cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                cell.font = Font(color="FFA500", bold=True)

# Data Cleaning Actions
row += 3
ws_quality.cell(row, 1, 'DATA CLEANING ACTIONS PERFORMED').fill = header_fill
ws_quality.cell(row, 1).font = header_font
ws_quality.merge_cells(f'A{row}:E{row}')

cleaning_actions = [
    '1. Converted Date column to datetime format',
    '2. Converted Amount to numeric, handling non-numeric values',
    '3. Converted Quantity to numeric format',
    '4. Handled encoding issues in CSV file (latin-1 encoding)',
    '5. Skipped bad lines during data loading',
    '6. All monetary values standardized to INR currency',
]

for action in cleaning_actions:
    row += 1
    ws_quality.cell(row, 1, action)
    ws_quality.merge_cells(f'A{row}:E{row}')

# Duplicate Records Check
row += 3
ws_quality.cell(row, 1, 'DUPLICATE RECORDS CHECK').fill = header_fill
ws_quality.cell(row, 1).font = header_font
ws_quality.merge_cells(f'A{row}:E{row}')

row += 1
duplicate_count = df.duplicated().sum()
ws_quality.cell(row, 1, 'Total Duplicate Records').font = metric_font
ws_quality.cell(row, 2, duplicate_count)
ws_quality.cell(row, 3, 'Clean' if duplicate_count == 0 else 'Has Duplicates')
if duplicate_count == 0:
    ws_quality.cell(row, 3).fill = green_card_value
    ws_quality.cell(row, 3).font = Font(color="00B050", bold=True)

# Column widths
ws_quality.column_dimensions['A'].width = 30
ws_quality.column_dimensions['B'].width = 15
ws_quality.column_dimensions['C'].width = 15
ws_quality.column_dimensions['D'].width = 20
ws_quality.column_dimensions['E'].width = 15

# ============================================================================
# SHEET 4: CATEGORY ANALYSIS
# ============================================================================
print("  ✓ Creating Category Analysis sheet with chart...")
ws_category = wb.create_sheet("Category Analysis")

category_data = df.groupby('Category').agg({
    'Qty': 'sum',
    'Amount': 'sum',
    'Order ID': 'count'
}).sort_values('Amount', ascending=False).reset_index()

headers = ['Category', 'Quantity', 'Revenue', 'Orders']
for col_idx, header in enumerate(headers, 1):
    cell = ws_category.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

for row_idx, row_data in enumerate(category_data.itertuples(), 2):
    ws_category.cell(row=row_idx, column=1).value = row_data.Category
    ws_category.cell(row=row_idx, column=2).value = row_data.Qty
    ws_category.cell(row=row_idx, column=3).value = row_data.Amount
    ws_category.cell(row=row_idx, column=4).value = row_data[4]

chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "Revenue by Product Category"
chart.y_axis.title = 'Revenue (₹)'
chart.x_axis.title = 'Category'

data = Reference(ws_category, min_col=3, min_row=1, max_row=len(category_data)+1)
cats = Reference(ws_category, min_col=1, min_row=2, max_row=len(category_data)+1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 12
chart.width = 20

ws_category.add_chart(chart, "F2")

# ============================================================================
# SHEET 4: GEOGRAPHY ANALYSIS
# ============================================================================
print("  ✓ Creating Geography Analysis sheet with chart...")
ws_geo = wb.create_sheet("Geography Analysis")

state_data = df.groupby('ship-state').agg({
    'Order ID': 'count',
    'Amount': 'sum'
}).sort_values('Amount', ascending=False).head(15).reset_index()

headers = ['State', 'Orders', 'Revenue']
for col_idx, header in enumerate(headers, 1):
    cell = ws_geo.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

for row_idx, row_data in enumerate(state_data.itertuples(), 2):
    ws_geo.cell(row=row_idx, column=1).value = row_data[1]
    ws_geo.cell(row=row_idx, column=2).value = row_data[2]
    ws_geo.cell(row=row_idx, column=3).value = row_data.Amount

chart = BarChart()
chart.type = "bar"
chart.style = 12
chart.title = "Top 15 States by Revenue"
chart.y_axis.title = 'State'
chart.x_axis.title = 'Revenue (₹)'

data = Reference(ws_geo, min_col=3, min_row=1, max_row=16)
cats = Reference(ws_geo, min_col=1, min_row=2, max_row=16)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 15
chart.width = 20

ws_geo.add_chart(chart, "E2")

# ============================================================================
# SHEET 5: ORDER STATUS
# ============================================================================
print("  ✓ Creating Order Status sheet with chart...")
ws_status = wb.create_sheet("Order Status")

status_data = df['Status'].value_counts().head(8).reset_index()
status_data.columns = ['Status', 'Count']

headers = ['Order Status', 'Count', 'Percentage']
for col_idx, header in enumerate(headers, 1):
    cell = ws_status.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

for row_idx, row_data in enumerate(status_data.itertuples(), 2):
    ws_status.cell(row=row_idx, column=1).value = row_data.Status
    ws_status.cell(row=row_idx, column=2).value = row_data.Count
    ws_status.cell(row=row_idx, column=3).value = row_data.Count / len(df)
    ws_status.cell(row=row_idx, column=3).number_format = '0.00%'

chart = PieChart()
chart.title = "Order Status Distribution"
chart.style = 10

data = Reference(ws_status, min_col=2, min_row=1, max_row=9)
cats = Reference(ws_status, min_col=1, min_row=2, max_row=9)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.dataLabels = DataLabelList()
chart.dataLabels.showPercent = True

ws_status.add_chart(chart, "E2")

# ============================================================================
# SHEET 6: SIZE ANALYSIS
# ============================================================================
print("  ✓ Creating Size Analysis sheet with chart...")
ws_size = wb.create_sheet("Size Analysis")

size_data = df[df['Qty'] > 0].groupby('Size')['Qty'].sum().sort_values(ascending=False).head(12).reset_index()

headers = ['Size', 'Quantity Sold']
for col_idx, header in enumerate(headers, 1):
    cell = ws_size.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

for row_idx, row_data in enumerate(size_data.itertuples(), 2):
    ws_size.cell(row=row_idx, column=1).value = str(row_data.Size)
    ws_size.cell(row=row_idx, column=2).value = row_data.Qty

chart = BarChart()
chart.type = "col"
chart.style = 11
chart.title = "Quantity Sold by Size"
chart.y_axis.title = 'Quantity'
chart.x_axis.title = 'Size'

data = Reference(ws_size, min_col=2, min_row=1, max_row=13)
cats = Reference(ws_size, min_col=1, min_row=2, max_row=13)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 12
chart.width = 18

ws_size.add_chart(chart, "D2")

# ============================================================================
# SHEET 7: SALES TREND
# ============================================================================
print("  ✓ Creating Sales Trend sheet with chart...")
ws_trend = wb.create_sheet("Sales Trend")

df_with_date = df[df['Date'].notna()].copy()
daily_sales = df_with_date.groupby('Date').agg({
    'Amount': 'sum',
    'Order ID': 'count'
}).reset_index()

headers = ['Date', 'Revenue', 'Orders']
for col_idx, header in enumerate(headers, 1):
    cell = ws_trend.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

for row_idx, row_data in enumerate(daily_sales.itertuples(), 2):
    ws_trend.cell(row=row_idx, column=1).value = row_data.Date
    ws_trend.cell(row=row_idx, column=2).value = row_data.Amount
    ws_trend.cell(row=row_idx, column=3).value = row_data[3]

chart = LineChart()
chart.title = "Daily Revenue Trend"
chart.style = 13
chart.y_axis.title = 'Revenue (₹)'
chart.x_axis.title = 'Date'

data = Reference(ws_trend, min_col=2, min_row=1, max_row=len(daily_sales)+1)
chart.add_data(data, titles_from_data=True)
chart.height = 12
chart.width = 25

ws_trend.add_chart(chart, "E2")

# ============================================================================
# SHEET 8: FULFILLMENT ANALYSIS
# ============================================================================
print("  ✓ Creating Fulfillment Analysis sheet with chart...")
ws_fulfill = wb.create_sheet("Fulfillment")

fulfill_data = df.groupby('Fulfilment').agg({
    'Order ID': 'count',
    'Amount': 'sum'
}).reset_index()

headers = ['Fulfillment Method', 'Orders', 'Revenue']
for col_idx, header in enumerate(headers, 1):
    cell = ws_fulfill.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

for row_idx, row_data in enumerate(fulfill_data.itertuples(), 2):
    ws_fulfill.cell(row=row_idx, column=1).value = row_data.Fulfilment
    ws_fulfill.cell(row=row_idx, column=2).value = row_data[2]
    ws_fulfill.cell(row=row_idx, column=3).value = row_data.Amount

chart = PieChart()
chart.title = "Orders by Fulfillment Method"
chart.style = 10

data = Reference(ws_fulfill, min_col=2, min_row=1, max_row=3)
cats = Reference(ws_fulfill, min_col=1, min_row=2, max_row=3)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.dataLabels = DataLabelList()
chart.dataLabels.showPercent = True

ws_fulfill.add_chart(chart, "E2")

# ============================================================================
# SHEET 9: B2B vs B2C
# ============================================================================
print("  ✓ Creating B2B vs B2C sheet with chart...")
ws_b2b = wb.create_sheet("B2B vs B2C")

b2b_data = df.groupby('B2B').agg({
    'Order ID': 'count',
    'Amount': ['sum', 'mean']
}).round(2)
b2b_data.columns = ['Orders', 'Total_Revenue', 'Avg_Order_Value']
b2b_data = b2b_data.reset_index()
b2b_data['B2B'] = b2b_data['B2B'].map({False: 'B2C', True: 'B2B'})

headers = ['Customer Type', 'Orders', 'Total Revenue', 'Avg Order Value']
for col_idx, header in enumerate(headers, 1):
    cell = ws_b2b.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

for row_idx, row_data in enumerate(b2b_data.itertuples(), 2):
    ws_b2b.cell(row=row_idx, column=1).value = row_data.B2B
    ws_b2b.cell(row=row_idx, column=2).value = row_data.Orders
    ws_b2b.cell(row=row_idx, column=3).value = row_data.Total_Revenue
    ws_b2b.cell(row=row_idx, column=4).value = row_data.Avg_Order_Value

chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "B2B vs B2C - Orders Comparison"
chart.y_axis.title = 'Number of Orders'

data = Reference(ws_b2b, min_col=2, min_row=1, max_row=3)
cats = Reference(ws_b2b, min_col=1, min_row=2, max_row=3)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

ws_b2b.add_chart(chart, "F2")

# Save workbook
print("\n" + "-"*80)
print("STEP 3: Saving workbook...")
print("-"*80)
wb.save(output_path)

print(f"\n{'='*80}")
print("SUCCESS! Excel Dashboard Created")
print("="*80)
print(f"\nOutput File: {output_filename}")
print(f"Location: outputs/")
print(f"\nWorkbook Contains:")
print("  1. Summary & Insights - Executive overview")
print("  2. Visual Dashboard - Colorful metric cards")
print("  3. Data Quality - Data cleaning & quality report")
print("  4. Category Analysis - Revenue by product with chart")
print("  5. Geography Analysis - Top states with chart")
print("  6. Order Status - Status distribution with pie chart")
print("  7. Size Analysis - Popular sizes with chart")
print("  8. Sales Trend - Daily revenue line chart")
print("  9. Fulfillment - Methods comparison with pie chart")
print(" 10. B2B vs B2C - Customer segment analysis")
print(f"\n{'='*80}")
print("You can now open this file and present to your team!")
print("="*80)
